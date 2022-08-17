import csv
import datetime
from openpyxl import Workbook, load_workbook


class ScriptImportation:
    def __init__(self, name_csv, xlsx_tlca_vide, xlsx_tlca_rempli, columns):
        # Fonction principale avec ces 4 paramètres
        # 1: Le nom du fichier CSV Moodle
        # 2: Le nom du fichier XLSX TLCA vide
        # 3: Le nom du fichier XLSX TLCA à remplir
        # 4: L'identifiant de la compétence à valider
        alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                    'U', 'V', 'W', 'X', 'Y', 'Z']

        wb = load_workbook(xlsx_tlca_vide)
        active = wb['Evaluations']

        with open(name_csv) as csv_file:
            csv_reader = csv.DictReader(csv_file, delimiter=";", quotechar='|')
            # ouvrir le fichier csv

            DATE_FORMAT = "%d %B %Y %H:%M %p"  # le format de date du csv

            temp = {}  # dict temporelle

            for row in csv_reader:  # loop pour parcourir le dict csv

                if (row['Grade/10.00'] == '-'):
                    temp[row['Surname'] + '.' + row['First name']] = ["-", row['Completed']]
                else:
                    if (row['Surname'] + '.' + row['First name'] in temp):
                        if (float(row['Grade/10.00']) > temp[row['Surname'] + '.' + row['First name']][0]):
                            temp[row['Surname'] + '.' + row['First name']] = [float(row['Grade/10.00']),
                                                                              row['Completed']]
                    else:
                        temp[row['Surname'] + '.' + row['First name']] = [float(row['Grade/10.00']), row['Completed']]

            x = 2

            # remplir le excel

            for item in temp:
                required_date = datetime.datetime.strptime(temp[item][1], DATE_FORMAT).strftime('%Y-%m-%d %H:%M') if (
                        temp[item][1] != '-') else '-'  # Conversion du bon format de date

                active['G' + str(x)] = str(round(temp[item][0] * 10)) + '%' if (
                            temp[item][1] != '-') else '-'  # Conversion en % et arrondisment
                active['F' + str(x)] = required_date
                active['E' + str(x)] = ('x' if (round(temp[item][0] * 10) >= 75) else ' ') if (
                            temp[item][1] != '-') else ' '  # Comparer si assez de % pour mettre un "X" dans la colonne

                for i in range(0, len(columns)):
                    active[alphabet[i + 7] + '1'] = columns[i]

                for column in columns:
                    for i in range(8, active.max_column + 1):
                        cell_obj = active.cell(row=1, column=i)
                        # print(cell_obj.value)
                        if column == cell_obj.value:
                            active[alphabet[i - 1] + str(x)] = ('x' if (round(temp[item][0] * 10) >= 75) else ' ') if (
                                        temp[item][1] != '-') else ' '

                x += 1

        wb.save(xlsx_tlca_rempli)  # sauvegarde du excel



